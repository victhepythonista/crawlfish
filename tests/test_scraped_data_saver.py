import os 
import json 
import openpyxl
import csv
from unittest import TestCase


from crawlfish.crawler.save import ScrapedDataSaver


test_scraped_data = [
			["NAME" , "AGE" , "EMAIL"],
			["John" , "44" , "john@gmail.com"],
			["Mao" , "66" , "maozedong@gmail.com"],
			["Heather" , "77" , "heather@gmail.com"],
			["Mark" , "34" , "mark@gmail.com"],

		]
test_scraped_data_dict = {
		"John":{"AGE":"44" , "EMAIL" : 'john@gmail.com'},
		"Mao":{"AGE":"66" , "EMAIL" : 'maozedong@gmail.com'},
		"Heather":{"AGE":"77" , "EMAIL" : 'heather@gmail.com'},
		"Mark":{"AGE":"34" , "EMAIL" : 'mark@gmail.com'},

}
test_scraped_data_json = json.dumps(test_scraped_data_dict)
data_saver = ScrapedDataSaver()

class TestScrapedDataSaver(TestCase):

	def test_saving_to_csv_file(self):
		test_file = "tests/test_scraped_data.csv"
		if os.path.isfile(test_file):
			os.remove(test_file)
		data_saver.to_csv_file(test_scraped_data , test_file)
		self.assertTrue(os.path.isfile(test_file))
		os.remove(test_file)
		# Test overwrite
			
		data_to_append = [["NAME" , "AGE" , "EMAIL"], ["Heisenberg" , "67" , "albrqq@gmail.com"],]
		expected_append_data = test_scraped_data + data_to_append[1:]
		data_saver.to_csv_file(test_scraped_data, test_file)
		data_saver.to_csv_file(data_to_append , test_file, overwrite = False)
		with open(test_file , 'r') as f:
			reader = csv.reader(f)
			rows = [row for row in reader ]
			self.assertTrue(rows == expected_append_data)


	def test_saving_to_spreadsheet_file(self):
		test_file = "tests/test_scraped_data.xlsx"
		if os.path.isfile(test_file):
			os.remove(test_file)
		data_saver.to_spreadsheet_file(test_scraped_data , test_file , sheet_name = "TestSheet1" , overwrite = True)
		self.assertTrue(os.path.isfile(test_file))
		workbook  = openpyxl.load_workbook(test_file)
		sheet = workbook["TestSheet1"]
		sheet_values = [ [ i for i in row] for row in sheet.values  ]
		self.assertTrue(sheet_values == test_scraped_data)
		os.remove(test_file)
		# write and then append data to first sheet and test data , test programatically and check manually 
		data_saver.to_spreadsheet_file(test_scraped_data , test_file , sheet_name = "TestSheet1" , overwrite = True)
		data_saver.to_spreadsheet_file(test_scraped_data , test_file , overwrite = False , sheet_name = "TestSheet1")
		workbook  = openpyxl.load_workbook(test_file)
		sheet = workbook["TestSheet1"]
		written_data = [ [ item for item in row ] for row in sheet.values ]# remove the tuples from sheet.values
		test_overwite_data = test_scraped_data + test_scraped_data[1:]
		self.assertTrue(written_data == test_overwite_data)
		# write to new sheet and test ovwrite ,  test programatically and check manually 
		data_saver.to_spreadsheet_file(test_scraped_data , test_file , sheet_name = "TestSheet2" )
		data_saver.to_spreadsheet_file(test_scraped_data[:-1] , test_file , sheet_name = "TestSheet2" )
		workbook = openpyxl.load_workbook(test_file)
		sheet = workbook["TestSheet2"]
		sheet_values = [ [ i for i in row] for row in sheet.values  ]
		self.assertTrue(sheet_values == test_scraped_data[:-1])
	
	def test_saving_to_dict(self):
		result = data_saver.to_dict(test_scraped_data , key_index = 0)
		self.assertTrue(result == test_scraped_data_dict)

	def test_saving_to_json(self):
		result = data_saver.to_json(test_scraped_data, key_index = 0 )
		self.assertTrue( result == test_scraped_data_json )

	def test_saving_to_json_file(self):
		test_file = "tests/data_json.json"
		if os.path.isfile(test_file):
			os.remove(test_file)
		result = data_saver.to_json_file(test_scraped_data, file = test_file )
		with open(test_file , "r" ) as f:
			data = json.load(f)
			self.assertTrue(json.dumps(data) == test_scraped_data_json)
		self.assertTrue( result == test_scraped_data_json )
		self.assertTrue(os.path.isfile(test_file))
		# test overwite 
		os.remove(test_file)
		json_data_to_append =  [
			["NAME" , "AGE" , "EMAIL"],
			["Heisenberg" , '67' , "albrqq@gmail.com"],]
		expected_append_data = test_scraped_data_dict | {
			 'Heisenberg': {'AGE': '67', 'EMAIL': 'albrqq@gmail.com'}}
		expected_append_data_json = json.dumps(expected_append_data)
		result = data_saver.to_json_file(test_scraped_data, file = test_file )
		append_result =  data_saver.to_json_file(json_data_to_append, file = test_file , overwrite = False )
		with open(test_file , "r" ) as f:
			data = json.load(f)
			self.assertTrue(expected_append_data_json == json.dumps(data) )
		os.remove(test_file)

