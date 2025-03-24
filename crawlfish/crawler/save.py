import os 
import csv
import json 
import openpyxl
from collections.abc import Iterable
from openpyxl import Workbook
from crawlfish.crawlfish_utils import check_file
from .save_options import SaveOption , InvalidSaveOptionError, CSVSaveOption , ExcelSaveOption , JSONSaveOption,JSONFileSaveOption



class ListDataSaver:
	'''
	This class contains methods for saving scraped data in different file formats

	'''
	@staticmethod
	def save_by_option(data:Iterable , save_option:SaveOption):
		'''Saves data by the specified option

		Parameters
		----------
		data:Iterable
			The data to e saved 
		save_option:SaveOption
			The save option instance to use 

		Returns
		-------
		data:[dict,list]
			The data saved
		'''
		if not isinstance(save_option , SaveOption):
			raise InvalidSaveOptionError( save_option)

		# CSV
		if isinstance(save_option ,CSVSaveOption ):
			return ListDataSaver.to_csv_file( data ,save_option.file , 
											save_option.overwrite , save_option.delimiter )
		# Excel spreadsheet
		if isinstance(save_option , ExcelSaveOption ):
			return ListDataSaver.to_spreadsheet_file(data , save_option.file , 
													save_option.sheet_name , save_option.overwrite)
		# JSON object 
		if isinstance(save_option , JSONSaveOption ):
			return ListDataSaver.to_json(data , save_option.key_index)
		# JSON file
		if isinstance(save_option , JSONFileSaveOption ):
			return ListDataSaver.to_json_file(data , save_option.key_index , save_option.file ,  save_option.overwrite ,
											 save_option.ensure_ascii  , save_option.indent  , 
											 save_option.encoding  )
		raise InvalidSaveOptionError(save_option)

	@staticmethod
	def to_csv_file(  data:Iterable , file:str = "Scraped_data.csv" , 
						overwrite:bool = True , delimiter:str="," ):
		'''Saves data to  CSV file 

		Parameters
		----------
		data:Iterable
			The data to be saved
		file:str
			The file where the data is to be stored
		overwrite:bool
			Whether or not to overwrite existing data in the file
		delimiter:str
			The delimiter in the csv file 


		Returns
		-------
		data 

		'''
		data = data 
		original_data = data
		data_to_write = []
		if not check_file(file):
			raise FileNotFoundError("Could not write to file - {}".format(file))
		# try and get existing data
		file_size = os.path.getsize(file)
		existing_data = []
		if file_size > 0:
			# check for existing data
			with open(file, 'r' ) as f:
				existing_data = [ row for row in csv.reader(f)]
		if overwrite:
			data_to_write = data
		else:
			data_to_write = existing_data + data[1:] 

		with open(file , "w" , newline= '') as f:
			writer = csv.writer(f , delimiter = delimiter)
			for row in data_to_write:
				writer.writerow(row)
		return original_data


	@staticmethod 
	def to_spreadsheet_file(data:Iterable , file:str , sheet_name = "data" , 
					overwrite = True):
		'''Saves data to an excel spreadheet file . 
		The program will not overwite existing data provide that a unique sheet name id provided
		
		Parameters
		----------
		data:Iterable
			The data to be saved
		file:str
			The file where the data is to be stored
		overwrite:bool
			Whether or not to overwrite existing data in the file
		sheet_name:str
			The name of the sheet in the excel file where the data is to be stored

		Returns
		-------
		data
		'''
		original_data = data 
		data = data 
		if not check_file(file):
			return FileNotFoundError("Could not create or write to file - {}".format(file))
		# try and get existing data 
		workbook = ""
		sheet = ""
		file_size = os.path.getsize(file)

		if file_size > 0:
			# file contains something maybe 
			try:
				workbook = openpyxl.load_workbook(file)
			except:
				raise 
		if not workbook:
			workbook = Workbook()
		if sheet_name in workbook.sheetnames:
			if overwrite:
				del workbook[sheet_name]
				sheet = workbook.create_sheet(sheet_name)
			else:
				# remove the header from data since we are now appending to existing data 
				if len(data) > 1:
					data = data[1:]
				else:
					data = []
				sheet = workbook[sheet_name]
		else:
			sheet = workbook.create_sheet(sheet_name)
		for row in data:
			sheet.append(row)
		workbook.save(file)
		return original_data 


	@staticmethod
	def to_dict(data:Iterable , key_index:int=0):
		'''Converts the data to a dict object
		
		Parameters
		----------
		data:Iterable
			The list to convert to a dict
		key_index:int
			The index of the item in each list in the data list that is to be used as a key 

		Returns
		-------
		data_in_dict_form:dict
			The data in dictionary form

		'''
		data_in_dict_form ={}
		headers = data[0]
		for row in data:
			if data.index(row) == 0:
				continue
			header = row[key_index]
			value = {}
			for item in row:
				index  = row.index(item)
				if index == key_index:
					continue
				print(index , headers)
				value[headers[index]] = item 
			data_in_dict_form[header] = value 
		return data_in_dict_form



	@staticmethod
	def to_json(data:Iterable , key_index = 0 ):
		'''Converts the data to a json object ,same as to_dict , WILL BE DEPRECATED

		Parameters
		----------
		data:Iterable
			The list to convert to a dict
		key_index:int
			The index of the item in each list in the data list that is to be used as a key 

		Returns
		-------
		data_in_dict_form:dict
			The data in dictionary form'''
		data_in_dict_form = ListDataSaver.to_dict(data , key_index = key_index)
		 
		return data_in_dict_form

	@staticmethod
	def to_json_file(data:Iterable ,
				 key_index = 0,
				  file:str = "scraped_data.json" ,
				 overwrite:bool = True,
				  ensure_ascii:bool = False  ,
				   indent:int = 4 ,
				    encoding:str = 'utf-8' ):
		'''Converts the data to a json object

		Parameters
		----------
		data:Iterable
			The list to convert to a dict
		key_index:int
			The index of the item in each list in the data list that is to be used as a key 
		file:str
			The name of the JSON file where the data will e stored
		overwrite:bool
			Whether or not to overwrite existing data in the file
		ensure_ascii:bool
			Whether or not to ensure ascii characters
		indent:int
			The indentation to be used in the file
		encoding:str
			The encoding to e used during writing to the file 
			
		Returns
		-------
		data_in_dict_form:dict
			The data in dictionary form
		'''
		if not check_file(file):
			raise FileNotFoundError("Could not load or write to file :( ")
		data_in_dict_form = ListDataSaver.to_dict(data , key_index = key_index)
		existing_json = ""
		try:
			with open(file , "r")  as f:
				existing_json = json.load(f)
		except (UnicodeDecodeError , json.JSONDecodeError) as e :
			pass 
		except:
			raise
		data_to_write = {}
		if overwrite:
			data_to_write = data_in_dict_form
		else:
			data_to_write = existing_json | data_in_dict_form
		with open(file, 'w', encoding=encoding) as f:
			json.dump(data_to_write, f, ensure_ascii=ensure_ascii, indent=indent)
		return data_in_dict_form
	